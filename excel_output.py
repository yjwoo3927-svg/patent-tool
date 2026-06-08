"""
특허 서지정보 → Excel 출력 모듈
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import os

# 색상 정의
COLORS = {
    "header_bg":    "1A3A5C",
    "header_font":  "FFFFFF",
    "kr_bg":        "E1F5EE",
    "us_bg":        "E6F1FB",
    "ep_bg":        "EEEDFE",
    "wo_bg":        "FAF0E6",
    "cn_bg":        "FFF3E0",
    "alt_row":      "F7F9FC",
    "white":        "FFFFFF",
    "accent":       "1D9E75",
    "title_bg":     "0D2137",
}

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COUNTRY_COLORS = {
    "KR": COLORS["kr_bg"],
    "US": COLORS["us_bg"],
    "EP": COLORS["ep_bg"],
    "WO": COLORS["wo_bg"],
    "CN": COLORS["cn_bg"],
}


def _apply_header(ws, row, cols, bg=COLORS["header_bg"], fg=COLORS["header_font"]):
    for col, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True, color=fg, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_patent_excel(patents: list, output_path: str) -> str:
    wb = openpyxl.Workbook()

    # ── Sheet 1: 서지정보 목록 ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "서지정보 목록"
    ws1.row_dimensions[1].height = 24
    ws1.freeze_panes = "A3"

    # 제목 행
    ws1.merge_cells("A1:M1")
    title_cell = ws1["A1"]
    title_cell.value = "특허 서지정보 수집 결과"
    title_cell.font = Font(bold=True, color=COLORS["header_font"], size=13)
    title_cell.fill = PatternFill("solid", fgColor=COLORS["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "No.", "출원번호/특허번호", "발명의 명칭", "출원인/권리자",
        "발명자", "IPC 코드", "출원일", "공개/등록일",
        "국가", "현황", "청구항 수", "피인용 수", "출처 DB"
    ]
    _apply_header(ws1, 2, headers)

    widths = [5, 18, 40, 20, 22, 18, 12, 12, 7, 10, 9, 9, 10]
    _set_col_widths(ws1, widths)

    for i, p in enumerate(patents, 1):
        row = i + 2
        country = p.get("country", "")
        row_bg = COUNTRY_COLORS.get(country, COLORS["white"]) if i % 2 == 0 else COLORS["white"]
        data = [
            i,
            p.get("patent_number", ""),
            p.get("title", ""),
            p.get("assignee", ""),
            p.get("inventors", ""),
            p.get("ipc_codes", ""),
            p.get("application_date", ""),
            p.get("publication_date", ""),
            country,
            p.get("status", ""),
            p.get("claims_count", 0),
            p.get("citations", 0),
            p.get("db_source", ""),
        ]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=row_bg)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col in [1, 9, 10, 11, 12, 13] else "left",
                wrap_text=(col == 3)
            )
            cell.font = Font(size=9)
        ws1.row_dimensions[row].height = 22

    # 합계 행
    total_row = len(patents) + 3
    ws1.cell(total_row, 1, "합계").font = Font(bold=True)
    ws1.cell(total_row, 11, f'=SUM(K3:K{total_row-1})').font = Font(bold=True)
    ws1.cell(total_row, 12, f'=SUM(L3:L{total_row-1})').font = Font(bold=True)
    for col in range(1, 14):
        ws1.cell(total_row, col).fill = PatternFill("solid", fgColor="E8EDF2")
        ws1.cell(total_row, col).border = THIN_BORDER

    # ── Sheet 2: 요약 통계 ────────────────────────────────────────────────
    ws2 = wb.create_sheet("요약 통계")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 30

    ws2.merge_cells("A1:D1")
    h = ws2["A1"]
    h.value = "특허 수집 요약 통계"
    h.font = Font(bold=True, color=COLORS["header_font"], size=12)
    h.fill = PatternFill("solid", fgColor=COLORS["title_bg"])
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    # 국가별 집계
    from collections import Counter
    country_cnt = Counter(p["country"] for p in patents)
    assignee_cnt = Counter(p["assignee"] for p in patents)
    db_cnt = Counter(p["db_source"] for p in patents)

    _apply_header(ws2, 2, ["국가", "건수", "비율(%)", "비고"])
    r = 3
    for country, cnt in sorted(country_cnt.items(), key=lambda x: -x[1]):
        ws2.cell(r, 1, country)
        ws2.cell(r, 2, cnt)
        ws2.cell(r, 3, f"=B{r}/SUM($B$3:$B${r+len(country_cnt)-1})*100")
        ws2.cell(r, 3).number_format = "0.0"
        for col in range(1, 5):
            ws2.cell(r, col).border = THIN_BORDER
            ws2.cell(r, col).alignment = Alignment(horizontal="center")
        r += 1
    total_r = r
    ws2.cell(r, 1, "합계").font = Font(bold=True)
    ws2.cell(r, 2, f"=SUM(B3:B{r-1})").font = Font(bold=True)
    for col in range(1, 5):
        ws2.cell(r, col).fill = PatternFill("solid", fgColor="E8EDF2")
        ws2.cell(r, col).border = THIN_BORDER

    # 차트: 국가별 건수
    chart_r = r + 2
    _apply_header(ws2, chart_r, ["출원인/권리자", "건수"])
    cr = chart_r + 1
    for assignee, cnt in assignee_cnt.most_common(10):
        ws2.cell(cr, 1, assignee)
        ws2.cell(cr, 2, cnt)
        for col in range(1, 3):
            ws2.cell(cr, col).border = THIN_BORDER
        cr += 1

    # Bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.title = "출원인별 특허 건수 Top 10"
    chart.y_axis.title = "건수"
    chart.x_axis.title = "출원인"
    chart.width = 20
    chart.height = 12
    data_ref = Reference(ws2, min_col=2, min_row=chart_r, max_row=cr - 1)
    cats_ref = Reference(ws2, min_col=1, min_row=chart_r + 1, max_row=cr - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "1D9E75"
    ws2.add_chart(chart, f"D{chart_r}")

    # Pie chart - 국가별
    pie = PieChart()
    pie.title = "국가별 특허 분포"
    pie.width = 14
    pie.height = 12
    pie_data = Reference(ws2, min_col=2, min_row=2, max_row=2 + len(country_cnt))
    pie_cats = Reference(ws2, min_col=1, min_row=3, max_row=2 + len(country_cnt))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    ws2.add_chart(pie, f"D{chart_r + 18}")

    # ── Sheet 3: 상세 내용 (초록 포함) ────────────────────────────────────
    ws3 = wb.create_sheet("상세 내용")
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 18

    ws3.merge_cells("A1:C1")
    h3 = ws3["A1"]
    h3.value = "특허 상세 서지정보 (초록 포함)"
    h3.font = Font(bold=True, color="FFFFFF", size=12)
    h3.fill = PatternFill("solid", fgColor=COLORS["title_bg"])
    h3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22

    for i, p in enumerate(patents):
        base = i * 9 + 2
        fields = [
            ("출원/특허번호", p.get("patent_number", "")),
            ("발명의 명칭", p.get("title", "")),
            ("출원인/권리자", p.get("assignee", "")),
            ("발명자", p.get("inventors", "")),
            ("IPC 코드", p.get("ipc_codes", "")),
            ("출원일 / 공개일", f"{p.get('application_date', '')} / {p.get('publication_date', '')}"),
            ("국가 / 현황", f"{p.get('country', '')} / {p.get('status', '')}"),
            ("초록", p.get("abstract", "")),
        ]
        # 구분선
        ws3.merge_cells(f"A{base}:C{base}")
        sep = ws3[f"A{base}"]
        sep.value = f"▶ 특허 {i+1}"
        sep.font = Font(bold=True, color="FFFFFF", size=10)
        sep.fill = PatternFill("solid", fgColor=COLORS["accent"])
        sep.alignment = Alignment(vertical="center")
        ws3.row_dimensions[base].height = 18

        for j, (label, val) in enumerate(fields):
            r = base + 1 + j
            lbl_cell = ws3.cell(r, 1, label)
            lbl_cell.font = Font(bold=True, size=9)
            lbl_cell.fill = PatternFill("solid", fgColor="F0F4F8")
            lbl_cell.alignment = Alignment(vertical="top")
            lbl_cell.border = THIN_BORDER

            ws3.merge_cells(f"B{r}:C{r}")
            val_cell = ws3.cell(r, 2, val)
            val_cell.font = Font(size=9)
            val_cell.alignment = Alignment(wrap_text=True, vertical="top")
            val_cell.border = THIN_BORDER
            ws3.row_dimensions[r].height = 30 if label == "초록" else 16

    wb.save(output_path)
    return output_path
